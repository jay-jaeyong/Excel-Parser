"""
Two-stage Excel parser.

Stage 1 — Structural (openpyxl):
  - Reads raw cells including merge info
  - Detects table boundaries by scanning for dense rectangular regions
  - Unmerges cells by propagating values
  - Detects multi-row headers via heuristics
  - Returns ParseResult with a confidence score

Stage 2 — LLM fallback (OpenAI / Gemini):
  - Triggered when Stage 1 confidence < threshold
  - Sends compact cell grid JSON to the LLM
  - LLM returns {headers, rows} JSON
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ─────────────────────────────────────────────────────────────────────────────
# Data Classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ParseResult:
    headers: list[str]
    rows: list[list[Any]]
    confidence: float
    stage: int  # 1 = heuristic, 2 = LLM
    warnings: list[str] = field(default_factory=list)
    raw_grid: Optional[list[list[Any]]] = field(default=None, repr=False)
    sheet_names: list[str] = field(default_factory=list)
    active_sheet: str = ""


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — Structural parser
# ─────────────────────────────────────────────────────────────────────────────

def _cell_value(cell) -> Any:
    """Return cleaned cell value."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _build_raw_grid(ws: Worksheet) -> tuple[list[list[Any]], dict[tuple[int, int], Any]]:
    """
    Read all cells into a 2-D list (row-major, 1-indexed offset to 0-indexed).
    Expand merged cells: the top-left cell keeps its value; all other cells
    in the merge get the same value propagated.
    Returns (grid, merge_map) where merge_map records which cells were synthetic.
    """
    if ws.max_row is None or ws.max_column is None:
        return [], {}

    # Build merge lookup: (row, col) → top-left value
    merge_values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        val = _cell_value(top_left)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_values[(row, col)] = val

    grid: list[list[Any]] = []
    for r in range(1, ws.max_row + 1):
        row_data: list[Any] = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in merge_values:
                row_data.append(merge_values[(r, c)])
            else:
                cell = ws.cell(r, c)
                row_data.append(_cell_value(cell))
        grid.append(row_data)

    return grid, merge_values


def _find_table_region(grid: list[list[Any]]) -> tuple[int, int, int, int]:
    """
    Detect the bounding box of the main table by finding the densest
    rectangular region of non-None values.
    Returns (start_row, start_col, end_row, end_col) — all 0-indexed.
    """
    if not grid:
        return 0, 0, 0, 0

    nrows = len(grid)
    ncols = max(len(r) for r in grid)

    # Row density: proportion of non-None cells per row
    row_density = []
    for row in grid:
        non_null = sum(1 for v in row if v is not None)
        row_density.append(non_null / ncols if ncols else 0)

    # Find contiguous block of rows with density > 0.2
    threshold = 0.2
    start_row, end_row = 0, nrows - 1
    for i, d in enumerate(row_density):
        if d >= threshold:
            start_row = i
            break
    for i in range(nrows - 1, -1, -1):
        if row_density[i] >= threshold:
            end_row = i
            break

    # Column density within detected row range
    col_density = []
    for c in range(ncols):
        non_null = sum(
            1 for r in range(start_row, end_row + 1)
            if c < len(grid[r]) and grid[r][c] is not None
        )
        total = end_row - start_row + 1
        col_density.append(non_null / total if total else 0)

    start_col, end_col = 0, ncols - 1
    for i, d in enumerate(col_density):
        if d >= threshold:
            start_col = i
            break
    for i in range(ncols - 1, -1, -1):
        if col_density[i] >= threshold:
            end_col = i
            break

    return start_row, start_col, end_row, end_col


def _detect_header_rows(subgrid: list[list[Any]]) -> int:
    """
    Heuristically determine how many top rows are headers.
    Clues: top rows contain more strings, subsequent rows contain more numbers.
    Returns number of header rows (1 or 2 for multi-level).
    """
    if len(subgrid) < 2:
        return 1

    def string_ratio(row: list[Any]) -> float:
        non_null = [v for v in row if v is not None]
        if not non_null:
            return 0.0
        return sum(1 for v in non_null if isinstance(v, str)) / len(non_null)

    row0_sr = string_ratio(subgrid[0])
    row1_sr = string_ratio(subgrid[1])

    # If both top rows are mostly strings and very different from row 2
    if len(subgrid) > 2:
        row2_sr = string_ratio(subgrid[2])
        if row0_sr > 0.7 and row1_sr > 0.7 and row2_sr < 0.5:
            return 2  # multi-level header

    return 1


def _flatten_headers(header_rows: list[list[Any]]) -> list[str]:
    """
    Flatten multi-level headers into single strings by joining with ' > '.
    Fills NoneS by carrying forward from previous column (common in merged headers).
    """
    if len(header_rows) == 1:
        row = header_rows[0]
        return [str(v) if v is not None else f"Column_{i+1}" for i, v in enumerate(row)]

    ncols = max(len(r) for r in header_rows)
    result: list[str] = []
    carry: list[Optional[str]] = [None] * ncols

    flattened: list[list[Optional[str]]] = []
    for row in header_rows:
        padded = list(row) + [None] * (ncols - len(row))
        # carry forward within this row for None cells (horizontal merge effect)
        last_val = None
        filled: list[Optional[str]] = []
        for v in padded:
            if v is not None:
                last_val = str(v)
            filled.append(last_val)
        flattened.append(filled)

    for c in range(ncols):
        parts = []
        seen: set[str] = set()
        for row in flattened:
            val = row[c]
            if val and val not in seen:
                parts.append(val)
                seen.add(val)
        result.append(" > ".join(parts) if parts else f"Column_{c+1}")

    return result


def _compute_confidence(
    grid: list[list[Any]],
    start_row: int, start_col: int, end_row: int, end_col: int,
    n_header_rows: int,
    merge_values: dict,
) -> float:
    """
    Compute confidence score (0.0–1.0) for Stage 1 result.
    Penalises:
    - Very small tables
    - High merge fraction (complex layouts)
    - Low data density
    """
    score = 1.0

    nrows = end_row - start_row + 1
    ncols = end_col - start_col + 1

    # Too small
    if nrows <= n_header_rows:
        return 0.1
    if ncols < 2:
        score -= 0.2

    # Data density
    data_rows = grid[start_row + n_header_rows: end_row + 1]
    total_cells = sum(ncols for _ in data_rows)
    non_null = sum(
        1 for row in data_rows
        for v in row[start_col: start_col + ncols]
        if v is not None
    )
    density = non_null / total_cells if total_cells else 0
    if density < 0.5:
        score -= 0.3
    elif density < 0.7:
        score -= 0.1

    # Merge complexity penalty
    merge_fraction = len(merge_values) / (nrows * ncols) if nrows * ncols else 0
    if merge_fraction > 0.4:
        score -= 0.3
    elif merge_fraction > 0.2:
        score -= 0.15

    return max(0.0, min(1.0, score))


def parse_stage1(wb: Workbook, sheet_name: Optional[str] = None) -> ParseResult:
    if sheet_name and sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
    grid, merge_values = _build_raw_grid(ws)

    sheet_names = wb.sheetnames
    active_sheet = ws.title

    if not grid:
        return ParseResult(headers=[], rows=[], confidence=0.0, stage=1,
                           warnings=["Empty worksheet"], raw_grid=grid,
                           sheet_names=sheet_names, active_sheet=active_sheet)

    sr, sc, er, ec = _find_table_region(grid)
    subgrid = [row[sc: ec + 1] for row in grid[sr: er + 1]]

    n_header_rows = _detect_header_rows(subgrid)
    header_rows = subgrid[:n_header_rows]
    data_rows = subgrid[n_header_rows:]

    headers = _flatten_headers(header_rows)
    # Normalise data rows to match header count
    ncols = len(headers)
    rows = [
        [(row[c] if c < len(row) else None) for c in range(ncols)]
        for row in data_rows
        if any(v is not None for v in row)
    ]

    confidence = _compute_confidence(grid, sr, sc, er, ec, n_header_rows, merge_values)
    warnings: list[str] = []
    if merge_values:
        warnings.append(f"Detected {len(ws.merged_cells.ranges)} merged cell region(s); values propagated.")

    return ParseResult(
        headers=headers,
        rows=rows,
        confidence=confidence,
        stage=1,
        warnings=warnings,
        raw_grid=grid,
        sheet_names=sheet_names,
        active_sheet=active_sheet,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — LLM fallback
# ─────────────────────────────────────────────────────────────────────────────

_SYSTEM_PROMPT = (
    "You are an expert at reading Excel spreadsheet data. "
    "Given a JSON representation of a cell grid (list of rows, each row is a list of cell values), "
    "identify the table headers and data rows. "
    "Respond ONLY with valid JSON in this exact format:\n"
    '{"headers": ["col1", "col2", ...], "rows": [[val, val, ...], ...]}\n'
    "Rules:\n"
    "- If headers span multiple rows, flatten them into one string per column using ' > '.\n"
    "- Skip empty rows and metadata rows above/below the table.\n"
    "- Preserve original data types (numbers as numbers, strings as strings).\n"
    "- Respond with ONLY the JSON object, no explanation."
)


def _grid_to_prompt(raw_grid: list[list[Any]]) -> str:
    # Limit to first 200 rows to stay within token limits
    truncated = raw_grid[:200]
    return json.dumps(truncated, default=str)


async def _call_openai(grid_json: str, settings) -> dict:
    try:
        from openai import AsyncOpenAI
    except ImportError:
        raise RuntimeError("openai package not installed. Add it to requirements.txt.")

    client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
    response = await client.chat.completions.create(
        model=settings.OPENAI_MODEL,
        messages=[
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user", "content": f"Cell grid:\n{grid_json}"},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )
    return json.loads(response.choices[0].message.content)


async def _call_gemini(grid_json: str, settings) -> dict:
    try:
        import google.generativeai as genai
    except ImportError:
        raise RuntimeError("google-generativeai package not installed. Add it to requirements.txt.")

    genai.configure(api_key=settings.GEMINI_API_KEY)
    model = genai.GenerativeModel(
        model_name=settings.GEMINI_MODEL,
        system_instruction=_SYSTEM_PROMPT,
    )
    prompt = f"Cell grid:\n{grid_json}"
    response = await model.generate_content_async(prompt)
    text = response.text.strip()
    # Strip markdown code fences if present
    text = re.sub(r"^```json\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


async def parse_stage2(stage1: ParseResult, settings) -> ParseResult:
    """Call LLM to re-interpret the raw grid."""
    if stage1.raw_grid is None:
        raise ValueError("No raw_grid available for Stage 2 parsing.")

    grid_json = _grid_to_prompt(stage1.raw_grid)

    try:
        if settings.LLM_PROVIDER == "openai":
            result = await _call_openai(grid_json, settings)
        elif settings.LLM_PROVIDER == "gemini":
            result = await _call_gemini(grid_json, settings)
        else:
            raise ValueError(f"Unknown LLM_PROVIDER: {settings.LLM_PROVIDER}")

        headers = [str(h) for h in result.get("headers", [])]
        rows = result.get("rows", [])

        return ParseResult(
            headers=headers,
            rows=rows,
            confidence=0.95,
            stage=2,
            warnings=[f"Stage 1 confidence was low; used {settings.LLM_PROVIDER.upper()} for parsing."],
        )
    except Exception as exc:
        # LLM failed — return Stage 1 result with a warning
        stage1.warnings.append(f"LLM fallback failed ({exc}); using Stage 1 result.")
        return stage1


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

async def parse_excel(file_bytes: bytes, settings, sheet_name: Optional[str] = None) -> ParseResult:
    """
    Main entry point.
    1. Run Stage 1 (structural parser).
    2. If confidence < threshold AND LLM is configured, run Stage 2.
    """
    import io
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,   # read computed values, not formulas
        read_only=False,  # need full access for merged_cells
    )

    result = parse_stage1(wb, sheet_name=sheet_name)

    # Cap rows at MAX_ROWS_PARSE to protect memory and downstream token limits
    max_rows = getattr(settings, "MAX_ROWS_PARSE", 100_000)
    if len(result.rows) > max_rows:
        result.rows = result.rows[:max_rows]
        result.warnings.append(
            f"파일에 행이 너무 많아 앞 {max_rows:,}행만 불러왔습니다. "
            f"AI 분석은 자동으로 대표 샘플을 사용합니다."
        )

    if result.confidence < settings.PARSER_CONFIDENCE_THRESHOLD and settings.llm_enabled:
        result = await parse_stage2(result, settings)

    return result
